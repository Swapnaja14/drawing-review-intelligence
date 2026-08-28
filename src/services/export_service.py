import json
import csv
import os
from typing import List, Dict, Any
from datetime import datetime

from src.core.dtos.export_dtos import ExportConfigDTO, ExportResultDTO, ExportFormat
from src.infrastructure.storage.repository import CommentRepository, ProjectRepository
from src.infrastructure.logging.logger import get_logger

logger = get_logger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl is not installed. Excel export will not work.")

class ExportService:
    def __init__(self, comment_repo: CommentRepository, project_repo: ProjectRepository):
        self.comment_repo = comment_repo
        self.project_repo = project_repo

    def export_drawing_comments(self, config: ExportConfigDTO) -> ExportResultDTO:
        try:
            if config.drawing_id:
                comments = self.comment_repo.get_comments_for_drawing(config.drawing_id)
            else:
                comments = []

            if config.filter_status:
                comments = [c for c in comments if c.get('status') == config.filter_status]

            if config.format == ExportFormat.EXCEL:
                return self._export_to_excel(comments, config)
            elif config.format == ExportFormat.JSON:
                return self._export_to_json(comments, config)
            elif config.format == ExportFormat.CSV:
                return self._export_to_csv(comments, config)
            else:
                return ExportResultDTO(
                    output_path=config.output_path,
                    format=config.format,
                    total_rows=0,
                    total_sheets=0,
                    file_size_bytes=0,
                    success=False,
                    error_message=f"Unsupported format: {config.format}"
                )
        except Exception as e:
            logger.error(f"Export failed: {str(e)}")
            return ExportResultDTO(
                output_path=config.output_path,
                format=config.format,
                total_rows=0,
                total_sheets=0,
                file_size_bytes=0,
                success=False,
                error_message=str(e)
            )

    def _export_to_excel(self, comments: List[Dict[str, Any]], config: ExportConfigDTO) -> ExportResultDTO:
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Error Tracker"

        headers = ['S.No', 'Drawing No', 'Page', 'Comment Text', 'Category', 'Confidence', 'Status', 'Reviewer', 'Timestamp']
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        status_colors = {
            'Approved': "00FF00",
            'Rejected': "FF0000",
            'Pending': "FFFF00",
            'Flagged': "FFA500"
        }

        for idx, comment in enumerate(comments, 1):
            row = [
                idx,
                comment.get('drawing_id', ''),
                comment.get('page_number', ''),
                comment.get('raw_text', ''),
                comment.get('category', ''),
                comment.get('confidence', 0.0) if config.include_confidence_scores else '',
                comment.get('status', 'Pending'),
                comment.get('reviewer_id', ''),
                comment.get('timestamp', '')
            ]
            ws.append(row)
            
            current_row = idx + 1
            status_val = comment.get('status', 'Pending')
            status_cell = ws.cell(row=current_row, column=7)
            if status_val in status_colors:
                status_cell.fill = PatternFill(start_color=status_colors[status_val], end_color=status_colors[status_val], fill_type="solid")

            confidence_cell = ws.cell(row=current_row, column=6)
            if config.include_confidence_scores:
                confidence_cell.number_format = '0.00%'
                
            for col_num in range(1, len(headers) + 1):
                ws.cell(row=current_row, column=col_num).border = thin_border

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2

        total_sheets = 1
        if config.include_summary_sheet:
            summary_ws = wb.create_sheet(title="Summary")
            summary_ws.append(['Metric', 'Value'])
            summary_ws.append(['Total Comments', len(comments)])
            
            statuses = [c.get('status', 'Pending') for c in comments]
            for stat in ['Approved', 'Rejected', 'Pending', 'Flagged']:
                summary_ws.append([f'{stat} Count', statuses.count(stat)])
                
            total_sheets += 1

        wb.save(config.output_path)
        file_size = os.path.getsize(config.output_path)

        return ExportResultDTO(
            output_path=config.output_path,
            format=config.format,
            total_rows=len(comments),
            total_sheets=total_sheets,
            file_size_bytes=file_size,
            success=True
        )

    def _export_to_json(self, comments: List[Dict[str, Any]], config: ExportConfigDTO) -> ExportResultDTO:
        data = {
            "metadata": {
                "export_date": datetime.now().isoformat(),
                "total_records": len(comments),
                "drawing_id": config.drawing_id
            },
            "comments": comments
        }
        
        with open(config.output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, default=str)
            
        file_size = os.path.getsize(config.output_path)

        return ExportResultDTO(
            output_path=config.output_path,
            format=config.format,
            total_rows=len(comments),
            total_sheets=1,
            file_size_bytes=file_size,
            success=True
        )

    def _export_to_csv(self, comments: List[Dict[str, Any]], config: ExportConfigDTO) -> ExportResultDTO:
        if not comments:
            with open(config.output_path, 'w', encoding='utf-8') as f:
                pass
            return ExportResultDTO(
                output_path=config.output_path,
                format=config.format,
                total_rows=0,
                total_sheets=1,
                file_size_bytes=0,
                success=True
            )

        headers = list(comments[0].keys())
        with open(config.output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(comments)
            
        file_size = os.path.getsize(config.output_path)

        return ExportResultDTO(
            output_path=config.output_path,
            format=config.format,
            total_rows=len(comments),
            total_sheets=1,
            file_size_bytes=file_size,
            success=True
        )
